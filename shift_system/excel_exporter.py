"""
Excel形式（.xlsx）でシフト表を出力する。
標準ライブラリ（zipfile + xml）のみで実装。openpyxl があれば自動的に高機能版を使用。
"""
import os
import io
import zipfile
from datetime import date, datetime
from typing import Dict, List

from .config import SHIFT_TYPES, WEEKDAY_LABELS, EXCEL_SETTINGS
from .models import Employee
from .scheduler import ShiftScheduler

# ---- openpyxl が使えるなら優先 ----
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _USE_OPENPYXL = True
except ImportError:
    _USE_OPENPYXL = False


def export_month(scheduler: ShiftScheduler, year: int, month: int, output_path: str) -> str:
    """指定年月のシフト表を Excel に出力する"""
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    if _USE_OPENPYXL:
        _export_openpyxl(scheduler, year, month, output_path)
    else:
        _export_native(scheduler, year, month, output_path)
    return output_path


# ==============================================================
# ネイティブ実装（標準ライブラリのみ）
# ==============================================================

# OOXML の共有文字列・スタイルなどを最小構成で構築

def _col_letter(n: int) -> str:
    """1-indexed 列番号 → A, B, ..., Z, AA, ..."""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def _xml_escape(s: str) -> str:
    return (str(s)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;"))


class _NativeXlsx:
    """zipfile で .xlsx を構築するミニマルクラス"""

    def __init__(self):
        self.sheets: List[dict] = []  # {name, rows: [[(value, style_idx), ...]]}
        self._styles: List[dict] = []  # style descriptor list
        self._style_cache: dict = {}
        self._strings: List[str] = []
        self._str_index: dict = {}
        # 列幅: sheet_idx -> {col_1indexed: width}
        self._col_widths: List[dict] = []
        # 行高: sheet_idx -> {row_1indexed: height}
        self._row_heights: List[dict] = []
        # freeze: sheet_idx -> "E4" etc
        self._freezes: List[str] = []
        # merged cells: sheet_idx -> list of "A1:C1"
        self._merges: List[List[str]] = []

    def add_sheet(self, name: str):
        idx = len(self.sheets)
        self.sheets.append({"name": name, "rows": {}})
        self._col_widths.append({})
        self._row_heights.append({})
        self._freezes.append(None)
        self._merges.append([])
        return idx

    def _get_str_idx(self, s: str) -> int:
        key = str(s)
        if key not in self._str_index:
            self._str_index[key] = len(self._strings)
            self._strings.append(key)
        return self._str_index[key]

    def _get_style_idx(self, fg=None, bold=False, font_color="000000",
                        font_size=11, halign="center", wrap=False) -> int:
        key = (fg, bold, font_color, font_size, halign, wrap)
        if key not in self._style_cache:
            self._style_cache[key] = len(self._styles)
            self._styles.append({
                "fg": fg, "bold": bold, "font_color": font_color,
                "font_size": font_size, "halign": halign, "wrap": wrap
            })
        return self._style_cache[key]

    def write_cell(self, sheet_idx: int, row: int, col: int, value,
                   fg=None, bold=False, font_color="000000",
                   font_size=11, halign="center", wrap=False):
        style = self._get_style_idx(fg, bold, font_color, font_size, halign, wrap)
        rows = self.sheets[sheet_idx]["rows"]
        if row not in rows:
            rows[row] = {}
        rows[row][col] = (value, style)

    def set_col_width(self, sheet_idx: int, col: int, width: float):
        self._col_widths[sheet_idx][col] = width

    def set_row_height(self, sheet_idx: int, row: int, height: float):
        self._row_heights[sheet_idx][row] = height

    def set_freeze(self, sheet_idx: int, cell: str):
        self._freezes[sheet_idx] = cell

    def add_merge(self, sheet_idx: int, merge_str: str):
        self._merges[sheet_idx].append(merge_str)

    # ---- XML builders ----

    def _build_content_types(self, n_sheets: int) -> bytes:
        parts = [
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">',
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>',
            '<Default Extension="xml" ContentType="application/xml"/>',
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>',
            '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>',
            '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>',
        ]
        for i in range(n_sheets):
            parts.append(f'<Override PartName="/xl/worksheets/sheet{i+1}.xml" '
                         f'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>')
        parts.append('</Types>')
        return "\n".join(parts).encode("utf-8")

    def _build_rels(self) -> bytes:
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            '</Relationships>'
        ).encode("utf-8")

    def _build_workbook_rels(self, n_sheets: int) -> bytes:
        parts = [
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">',
        ]
        for i in range(n_sheets):
            parts.append(f'<Relationship Id="rId{i+1}" '
                         f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
                         f'Target="worksheets/sheet{i+1}.xml"/>')
        parts.append(f'<Relationship Id="rId{n_sheets+1}" '
                     f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" '
                     f'Target="sharedStrings.xml"/>')
        parts.append(f'<Relationship Id="rId{n_sheets+2}" '
                     f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" '
                     f'Target="styles.xml"/>')
        parts.append('</Relationships>')
        return "\n".join(parts).encode("utf-8")

    def _build_workbook(self) -> bytes:
        parts = [
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">',
            '<sheets>',
        ]
        for i, sheet in enumerate(self.sheets):
            parts.append(f'<sheet name="{_xml_escape(sheet["name"])}" sheetId="{i+1}" r:id="rId{i+1}"/>')
        parts.append('</sheets></workbook>')
        return "\n".join(parts).encode("utf-8")

    def _build_shared_strings(self) -> bytes:
        parts = [
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
            f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            f'count="{len(self._strings)}" uniqueCount="{len(self._strings)}">',
        ]
        for s in self._strings:
            escaped = _xml_escape(s).replace("\n", "&#10;")
            parts.append(f'<si><t xml:space="preserve">{escaped}</t></si>')
        parts.append('</sst>')
        return "\n".join(parts).encode("utf-8")

    def _build_styles(self) -> bytes:
        # fonts
        fonts = ['<fonts count="{}">'.format(len(self._styles) + 1)]
        fonts.append('<font><sz val="11"/><name val="Calibri"/></font>')  # default
        for s in self._styles:
            bold_tag = "<b/>" if s["bold"] else ""
            color_tag = f'<color rgb="FF{s["font_color"]}"/>'
            fonts.append(f'<font>{bold_tag}<sz val="{s["font_size"]}"/>'
                         f'{color_tag}<name val="Calibri"/></font>')
        fonts.append('</fonts>')

        # fills (index 0,1 reserved)
        fills = ['<fills count="{}">'.format(len(self._styles) + 2)]
        fills.append('<fill><patternFill patternType="none"/></fill>')
        fills.append('<fill><patternFill patternType="gray125"/></fill>')
        for s in self._styles:
            if s["fg"]:
                fills.append(f'<fill><patternFill patternType="solid">'
                              f'<fgColor rgb="FF{s["fg"]}"/></patternFill></fill>')
            else:
                fills.append('<fill><patternFill patternType="none"/></fill>')
        fills.append('</fills>')

        # borders
        borders = ['<borders count="2">',
                   '<border><left/><right/><top/><bottom/><diagonal/></border>',
                   '<border><left style="thin"><color rgb="FFAAAAAA"/></left>'
                   '<right style="thin"><color rgb="FFAAAAAA"/></right>'
                   '<top style="thin"><color rgb="FFAAAAAA"/></top>'
                   '<bottom style="thin"><color rgb="FFAAAAAA"/></bottom>'
                   '<diagonal/></border>',
                   '</borders>']

        # cellStyleXfs
        xfs_base = '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'

        # cellXfs
        xfs = ['<cellXfs count="{}">'.format(len(self._styles) + 1)]
        xfs.append('<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>')
        for i, s in enumerate(self._styles):
            align_tag = f'<alignment horizontal="{s["halign"]}" vertical="center"'
            if s["wrap"]:
                align_tag += ' wrapText="1"'
            align_tag += '/>'
            xfs.append(
                f'<xf numFmtId="0" fontId="{i+1}" fillId="{i+2}" borderId="1" xfId="0" '
                f'applyFont="1" applyFill="1" applyBorder="1" applyAlignment="1">'
                f'{align_tag}</xf>'
            )
        xfs.append('</cellXfs>')

        xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            + "\n".join(fonts) + "\n".join(fills) + "\n".join(borders)
            + xfs_base + "\n".join(xfs)
            + '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
            + '</styleSheet>'
        )
        return xml.encode("utf-8")

    def _build_worksheet(self, sheet_idx: int) -> bytes:
        sheet = self.sheets[sheet_idx]
        rows_data = sheet["rows"]
        col_widths = self._col_widths[sheet_idx]
        row_heights = self._row_heights[sheet_idx]
        freeze = self._freezes[sheet_idx]
        merges = self._merges[sheet_idx]

        parts = [
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
        ]

        # column widths
        if col_widths:
            parts.append('<cols>')
            for col, w in sorted(col_widths.items()):
                parts.append(f'<col min="{col}" max="{col}" width="{w}" customWidth="1"/>')
            parts.append('</cols>')

        # sheetView (freeze)
        if freeze:
            col_str = ''.join(c for c in freeze if c.isalpha())
            row_num = ''.join(c for c in freeze if c.isdigit())
            col_idx = sum((ord(c) - 64) * (26 ** i) for i, c in enumerate(reversed(col_str)))
            row_idx = int(row_num)
            parts.append(
                '<sheetViews><sheetView workbookViewId="0">'
                f'<pane xSplit="{col_idx - 1}" ySplit="{row_idx - 1}" '
                f'topLeftCell="{freeze}" activePane="bottomRight" state="frozen"/>'
                '</sheetView></sheetViews>'
            )

        # sheetData
        parts.append('<sheetData>')
        for row_num in sorted(rows_data.keys()):
            ht = row_heights.get(row_num, "")
            ht_attr = f' ht="{ht}" customHeight="1"' if ht else ""
            parts.append(f'<row r="{row_num}"{ht_attr}>')
            for col_num in sorted(rows_data[row_num].keys()):
                value, style_idx = rows_data[row_num][col_num]
                cell_ref = f"{_col_letter(col_num)}{row_num}"
                real_style = style_idx + 1  # 0 is default
                if value is None or value == "":
                    parts.append(f'<c r="{cell_ref}" s="{real_style}"/>')
                elif isinstance(value, (int, float)):
                    parts.append(f'<c r="{cell_ref}" s="{real_style}" t="n"><v>{value}</v></c>')
                else:
                    si = self._get_str_idx(str(value))
                    parts.append(f'<c r="{cell_ref}" s="{real_style}" t="s"><v>{si}</v></c>')
            parts.append('</row>')
        parts.append('</sheetData>')

        # merges
        if merges:
            parts.append(f'<mergeCells count="{len(merges)}">'),
            for m in merges:
                parts.append(f'<mergeCell ref="{m}"/>')
            parts.append('</mergeCells>')

        parts.append('</worksheet>')
        return "\n".join(parts).encode("utf-8")

    def save(self, path: str):
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("[Content_Types].xml", self._build_content_types(len(self.sheets)))
            zf.writestr("_rels/.rels", self._build_rels())
            zf.writestr("xl/workbook.xml", self._build_workbook())
            zf.writestr("xl/_rels/workbook.xml.rels", self._build_workbook_rels(len(self.sheets)))
            zf.writestr("xl/sharedStrings.xml", self._build_shared_strings())
            zf.writestr("xl/styles.xml", self._build_styles())
            for i in range(len(self.sheets)):
                zf.writestr(f"xl/worksheets/sheet{i+1}.xml", self._build_worksheet(i))
        with open(path, "wb") as f:
            f.write(buf.getvalue())


# ==============================================================
# シフト表書き込み（ネイティブ版）
# ==============================================================

def _export_native(scheduler: ShiftScheduler, year: int, month: int, output_path: str):
    wb = _NativeXlsx()
    _write_shift_sheet_native(wb, scheduler, year, month)
    _write_summary_sheet_native(wb, scheduler, year, month)
    wb.save(output_path)


def _write_shift_sheet_native(wb: _NativeXlsx, scheduler: ShiftScheduler, year: int, month: int):
    si = wb.add_sheet(EXCEL_SETTINGS["sheet_name"])
    dates = scheduler.get_month_dates(year, month)
    matrix = scheduler.get_month_matrix(year, month)
    employees = list(scheduler.employees.values())
    today_str = date.today().isoformat()

    HDR = EXCEL_SETTINGS["header_color"]
    HDR_FC = EXCEL_SETTINGS["header_font_color"]

    def hcell(r, c, v, fg=HDR, fc=HDR_FC, bold=True, size=11):
        wb.write_cell(si, r, c, v, fg=fg, bold=bold, font_color=fc, font_size=size)

    def dcell(r, c, v, fg=None, bold=False, size=9):
        wb.write_cell(si, r, c, v, fg=fg, bold=bold, font_size=size)

    # タイトル
    hcell(1, 1, f"{year}年{month}月 シフト表", size=14)
    wb.add_merge(si, f"A1:{_col_letter(4 + len(dates))}1")
    wb.set_row_height(si, 1, 25)

    # 固定ヘッダー
    for col, label in enumerate(["No", "氏名", "役職", "雇用形態"], 1):
        hcell(2, col, label)
        hcell(3, col, "")

    # 日付ヘッダー
    for i, d_str in enumerate(dates):
        col = 5 + i
        d = datetime.strptime(d_str, "%Y-%m-%d").date()
        wd = d.weekday()
        day_num = str(d.day)
        wd_label = WEEKDAY_LABELS[wd]

        if wd == 5:
            bg = "BDD7EE"
        elif wd == 6:
            bg = "FCE4D6"
        else:
            bg = HDR

        fc = HDR_FC if bg == HDR else "000000"
        if d_str == today_str:
            bg = EXCEL_SETTINGS["today_color"]
            fc = "000000"

        hcell(2, col, day_num, fg=bg, fc=fc, size=9)
        hcell(3, col, wd_label, fg=bg, fc=fc, size=9)
        wb.set_col_width(si, col, 4.5)

    # 従業員行
    rest_shifts = {"休み", "有休"}
    for ri, emp in enumerate(employees):
        row = 4 + ri
        row_bg = "FFFFFF" if ri % 2 == 0 else "F2F2F2"
        for col, val in enumerate([ri + 1, emp.name, emp.role, emp.employment_type], 1):
            wb.write_cell(si, row, col, val, fg=row_bg, font_size=10)

        for i, d_str in enumerate(dates):
            col = 5 + i
            entry = scheduler.get_shift(emp.id, d_str)
            stype = matrix[emp.id][d_str]
            memo = entry.memo if entry else ""
            val = f"{stype}\n{memo}" if memo else stype
            bg = SHIFT_TYPES.get(stype, {}).get("color", "FFFFFF")
            dcell(row, col, val, fg=bg, size=9)

        # 出勤日数
        work = sum(1 for d in dates if matrix[emp.id][d] not in rest_shifts)
        summary_col = 5 + len(dates)
        wb.write_cell(si, row, summary_col, work, fg="FFFFFF", bold=True, font_size=10)

    # 集計列ヘッダー
    summary_col = 5 + len(dates)
    hcell(2, summary_col, "出勤日数")
    hcell(3, summary_col, "")
    wb.set_col_width(si, summary_col, 8)

    # 凡例
    legend_row = 4 + len(employees) + 2
    wb.write_cell(si, legend_row, 1, "【凡例】", bold=True, font_size=10)
    for li, (stype, info) in enumerate(SHIFT_TYPES.items()):
        col = 2 + li
        wb.write_cell(si, legend_row, col, stype, fg=info["color"], font_size=9)
        time_str = f"{info['start']}~{info['end']}" if info["start"] else "-"
        wb.write_cell(si, legend_row + 1, col, time_str, fg=info["color"], font_size=8)

    # 固定列幅
    wb.set_col_width(si, 1, 5)
    wb.set_col_width(si, 2, 12)
    wb.set_col_width(si, 3, 10)
    wb.set_col_width(si, 4, 10)
    wb.set_row_height(si, 2, 20)
    wb.set_row_height(si, 3, 15)
    wb.set_freeze(si, "E4")


def _write_summary_sheet_native(wb: _NativeXlsx, scheduler: ShiftScheduler, year: int, month: int):
    si = wb.add_sheet("月次集計")
    employees = list(scheduler.employees.values())
    summary = scheduler.summarize_month(year, month)
    shift_names = list(SHIFT_TYPES.keys())
    rest_shifts = {"休み", "有休"}

    HDR = EXCEL_SETTINGS["header_color"]
    HDR_FC = EXCEL_SETTINGS["header_font_color"]

    headers = ["No", "氏名", "役職", "雇用形態"] + shift_names + ["合計勤務日", "合計休日"]
    for col, h in enumerate(headers, 1):
        wb.write_cell(si, 1, col, h, fg=HDR, bold=True, font_color=HDR_FC, font_size=10)
        wb.set_col_width(si, col, 9)
    wb.set_col_width(si, 2, 14)

    for ri, emp in enumerate(employees):
        row = 2 + ri
        bg = "FFFFFF" if ri % 2 == 0 else "F2F2F2"
        emp_summary = summary.get(emp.id, {})
        for col, val in enumerate([ri + 1, emp.name, emp.role, emp.employment_type], 1):
            wb.write_cell(si, row, col, val, fg=bg, font_size=10)

        total_work = total_rest = 0
        for i, stype in enumerate(shift_names):
            cnt = emp_summary.get(stype, 0)
            col = 5 + i
            bg_c = SHIFT_TYPES[stype]["color"]
            wb.write_cell(si, row, col, cnt, fg=bg_c, font_size=10)
            if stype not in rest_shifts:
                total_work += cnt
            else:
                total_rest += cnt

        wb.write_cell(si, row, 5 + len(shift_names), total_work, bold=True, font_size=10)
        wb.write_cell(si, row, 6 + len(shift_names), total_rest, font_size=10)

    wb.set_freeze(si, "E2")


# ==============================================================
# openpyxl版（利用可能な場合）— 元の実装をここに残す
# ==============================================================

def _export_openpyxl(scheduler: ShiftScheduler, year: int, month: int, output_path: str):
    wb = Workbook()
    _write_shift_sheet_opx(wb, scheduler, year, month)
    _write_summary_sheet_opx(wb, scheduler, year, month)
    wb.save(output_path)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font_opx(bold=False, color="000000", size=11):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _border_opx():
    thin = Side(style="thin", color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center_opx(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _write_shift_sheet_opx(wb, scheduler, year, month):
    ws = wb.active
    ws.title = EXCEL_SETTINGS["sheet_name"]
    dates = scheduler.get_month_dates(year, month)
    matrix = scheduler.get_month_matrix(year, month)
    employees = list(scheduler.employees.values())
    today_str = date.today().isoformat()
    HDR = EXCEL_SETTINGS["header_color"]
    HDR_FC = EXCEL_SETTINGS["header_font_color"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + len(dates))
    tc = ws.cell(1, 1, f"{year}年{month}月 シフト表")
    tc.font = _font_opx(bold=True, size=14)
    tc.alignment = _center_opx()

    for col, label in enumerate(["No", "氏名", "役職", "雇用形態"], 1):
        for r in [2, 3]:
            c = ws.cell(r, col, label if r == 2 else "")
            c.fill = _fill(HDR); c.font = _font_opx(bold=True, color=HDR_FC)
            c.alignment = _center_opx(); c.border = _border_opx()

    for i, d_str in enumerate(dates):
        col = 5 + i
        d = datetime.strptime(d_str, "%Y-%m-%d").date()
        wd = d.weekday()
        bg = "BDD7EE" if wd == 5 else ("FCE4D6" if wd == 6 else HDR)
        fc = "000000" if wd in (5, 6) else HDR_FC
        if d_str == today_str:
            bg = EXCEL_SETTINGS["today_color"]; fc = "000000"
        for r, v in [(2, str(d.day)), (3, WEEKDAY_LABELS[wd])]:
            c = ws.cell(r, col, v)
            c.fill = _fill(bg); c.font = _font_opx(bold=True, color=fc, size=9)
            c.alignment = _center_opx(); c.border = _border_opx()
        ws.column_dimensions[get_column_letter(col)].width = 4.5

    rest_shifts = {"休み", "有休"}
    for ri, emp in enumerate(employees):
        row = 4 + ri
        rb = "FFFFFF" if ri % 2 == 0 else "F2F2F2"
        for col, val in enumerate([ri+1, emp.name, emp.role, emp.employment_type], 1):
            c = ws.cell(row, col, val)
            c.fill = _fill(rb); c.font = _font_opx(size=10)
            c.alignment = _center_opx(); c.border = _border_opx()
        for i, d_str in enumerate(dates):
            entry = scheduler.get_shift(emp.id, d_str)
            stype = matrix[emp.id][d_str]
            memo = entry.memo if entry else ""
            c = ws.cell(row, 5+i, f"{stype}\n{memo}" if memo else stype)
            c.fill = _fill(SHIFT_TYPES.get(stype, {}).get("color", "FFFFFF"))
            c.font = _font_opx(size=9); c.alignment = _center_opx(wrap=True)
            c.border = _border_opx()
        work = sum(1 for d in dates if matrix[emp.id][d] not in rest_shifts)
        sc = ws.cell(row, 5+len(dates), work)
        sc.font = _font_opx(bold=True); sc.alignment = _center_opx(); sc.border = _border_opx()

    sc_h = ws.cell(2, 5+len(dates), "出勤日数")
    sc_h.fill = _fill(HDR); sc_h.font = _font_opx(bold=True, color=HDR_FC)
    sc_h.alignment = _center_opx(); sc_h.border = _border_opx()
    ws.cell(3, 5+len(dates), "").fill = _fill(HDR)

    lr = 4 + len(employees) + 2
    ws.cell(lr, 1, "【凡例】").font = _font_opx(bold=True)
    for li, (stype, info) in enumerate(SHIFT_TYPES.items()):
        c = ws.cell(lr, 2+li, stype)
        c.fill = _fill(info["color"]); c.alignment = _center_opx(); c.border = _border_opx()
        ts = f"{info['start']}~{info['end']}" if info["start"] else "—"
        ws.cell(lr+1, 2+li, ts).alignment = _center_opx()

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions[get_column_letter(5+len(dates))].width = 8
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "E4"


def _write_summary_sheet_opx(wb, scheduler, year, month):
    ws = wb.create_sheet("月次集計")
    employees = list(scheduler.employees.values())
    summary = scheduler.summarize_month(year, month)
    shift_names = list(SHIFT_TYPES.keys())
    rest_shifts = {"休み", "有休"}
    HDR = EXCEL_SETTINGS["header_color"]
    HDR_FC = EXCEL_SETTINGS["header_font_color"]
    headers = ["No", "氏名", "役職", "雇用形態"] + shift_names + ["合計勤務日", "合計休日"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.fill = _fill(HDR); c.font = _font_opx(bold=True, color=HDR_FC)
        c.alignment = _center_opx(); c.border = _border_opx()
    for ri, emp in enumerate(employees):
        row = 2 + ri
        bg = "FFFFFF" if ri % 2 == 0 else "F2F2F2"
        es = summary.get(emp.id, {})
        for col, val in enumerate([ri+1, emp.name, emp.role, emp.employment_type], 1):
            c = ws.cell(row, col, val)
            c.fill = _fill(bg); c.font = _font_opx(size=10)
            c.alignment = _center_opx(); c.border = _border_opx()
        tw = tr = 0
        for i, stype in enumerate(shift_names):
            cnt = es.get(stype, 0)
            c = ws.cell(row, 5+i, cnt)
            c.fill = _fill(SHIFT_TYPES[stype]["color"])
            c.alignment = _center_opx(); c.border = _border_opx()
            (tr if stype in rest_shifts else tw)  # dummy
            if stype in rest_shifts:
                tr += cnt
            else:
                tw += cnt
        ws.cell(row, 5+len(shift_names), tw).font = _font_opx(bold=True)
        ws.cell(row, 6+len(shift_names), tr)
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 9
    ws.column_dimensions["B"].width = 14
    ws.freeze_panes = "E2"
